from openpyxl import Workbook, load_workbook
import random


def get_profile_url(name, first_name, last_name, user_list):
    candidate_list = []
    for mem in user_list['members']:
        if name in mem['profile']['real_name_normalized']:
            candidate_list.append(mem)

    if len(candidate_list) < 1:
        print(candidate_list)
        for mem in user_list['members']:
            if hasattr(mem['profile'], 'first_name'):
                if first_name in mem['profile']['first_name'] and last_name in mem['profile']['last_name']:
                    candidate_list.append(mem)
    elif len(candidate_list) > 1:
        print(candidate_list)
        for c in candidate_list:
            if first_name in c['profile']['first_name'] and last_name in c['profile']['last_name']:
                return c['profile']['image_512']
    return candidate_list[0]['profile']['image_512'] if len(candidate_list) == 1 else "None"


def random_num(max_num):
    return random.randint(1, max_num)


class Utils:

    def get_info_by_index_num(workbook, index, prize_name, user_list):
        sheet = workbook['인원명부 공유용']
        num = 4 + random_num(index)

        while(sheet.cell(row=num, column=16).value == "None"):
            num = 4 + random_num(index)

        name = sheet.cell(row=num, column=2).value
        first_name = sheet.cell(row=num, column=3).value
        last_name = sheet.cell(row=num, column=4).value
        group = sheet.cell(row=num, column=7).value
        division = sheet.cell(row=num, column=8).value
        department = sheet.cell(row=num, column=9).value
        team = sheet.cell(row=num, column=10).value
        url = get_profile_url(name, first_name, last_name, user_list)
        dict = {
            "first_name": first_name,
            "last_name": last_name,
            "group": group,
            "division": division,
            "department": department,
            "team": team,
            "image_url": url
        }
        sheet['P' + str(num)] = prize_name
        return workbook, dict

    def reset_sheet(workbook):
        sheet = workbook['인원명부 공유용']
        sheet.delete_cols(16)
        return workbook
