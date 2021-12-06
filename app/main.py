from typing import Optional

from fastapi import FastAPI
from openpyxl.reader.excel import load_workbook
from pydantic import BaseModel
from slack_sdk import WebClient
from starlette.responses import JSONResponse
from util import Utils

app = FastAPI()

Slack_token = "SECRET"
client = WebClient(token=Slack_token)
user_list = client.users_list()
data_xlsx_ = './data.xlsx'
MAX_PEOPLE = 263


@app.get("/")
def read_root():
    for mem in user_list['members']:
        profile = mem['profile']
        if 'David' in profile['real_name']:
            print(profile)

    return "HELLO"


@ app.get("/pick/{prize_name}")
def pick_man(prize_name: str):
    workbook = load_workbook(data_xlsx_, data_only=True)
    wb, dict = Utils.get_info_by_index_num(
        workbook=workbook,
        index=MAX_PEOPLE,
        prize_name=prize_name,
        user_list=user_list,
    )
    wb.save(data_xlsx_)
    return JSONResponse(content=dict)


@ app.get("/reset/")
def resetting_data():
    workbook = load_workbook(data_xlsx_, data_only=True)
    wb = Utils.reset_sheet(workbook)
    wb.save(data_xlsx_)
    return "anything"
